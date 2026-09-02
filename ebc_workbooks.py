"""Build the Excel workbooks - one per group of trials that stands on its own.

    python ebc_workbooks.py <config.json>

Each workbook carries its own read-me: the protocol, how the numbers were produced and
what not to over-read, so a sheet that leaves this folder still explains itself.
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import json, numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, LineChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.drawing.image import Image as XLImage

import ebc_config as C
from ebc_paths import BASE, work_dir, out_dir

CFG = C.load(sys.argv[1] if len(sys.argv) > 1 else None)
WORK, OUT = work_dir(CFG), out_dir(CFG)
M = json.load(open(os.path.join(WORK, "merged.json"), encoding="utf-8"))
ALLROWS = json.load(open(os.path.join(WORK, "merged_rows.json"), encoding="utf-8"))
NOM = C.fill(M["protocol"])
CHECKS = M.get("checks", {})
TAG_ORDER = [r["tag"] for r in CFG["recordings"]]
SESSMETA = {s["tag"]: s for s in M["sessions"]}
REC_FILE = {r["tag"]: r["file"] for r in CFG["recordings"]}

# the sheet code below speaks in these older names; keep them as the bridge
for _r in ALLROWS:
    _r["block_kind"] = _r["role"]
    _r["gidx"] = _r["group_index"]
    _r["cs_onset_block_s"] = _r.get("session_clock_s")
    _r["cs_duration_measured_ms"] = _r.get("cs_duration_ms")

DES = C.design(NOM)
# Which latencies were counted as a conditioned response, read back from the run rather
# than recomputed here: when the study has a US-only baseline the window is measured from
# it, so no arithmetic on the protocol alone could reproduce it.  A merged.json written
# before the window was measured falls back to the protocol, which is how it was scored.
WIN = M.get("cr_window") or C.cr_window(NOM)
CR_LO, CR_HI = float(WIN["lo_ms"]), float(WIN["hi_ms"])
ALPHA = CR_LO
# The response labels have to be the exact strings ebc_score.classify() wrote into the
# rows, so every one of them comes out of the same window rather than being spelled out.
CR_LBL, UR_LBL, ALPHA_LBL = WIN["cr_label"], WIN["ur_label"], WIN["alpha_label"]
MOVING_LBL = WIN["moving_label"]
# what ebc_score.classify() emits for a recording anchored on the US instead of the CS
UR_PUFF_LBL, ALPHA_US_LBL = WIN["ur_puff_label"], WIN["alpha_us_label"]

# Where every trial sits in its own recording, for the worklist of trials somebody now
# has to look at.  merged.json already worked these out for the app and the CSV; reading
# them back means the three lists can never disagree.
MANUAL = {(t["session"], t["session_trial"]): t
          for t in (M.get("manual_review") or {}).get("trials", [])}


def clock(seconds):
    """A position in its own recording, in the form a video player shows."""
    if seconds is None:
        return ""
    mm, sec = divmod(float(seconds), 60.0)
    return "%d:%06.3f" % (int(mm), sec)


def at_in_video(r):
    t = MANUAL.get((r["session"], r["session_trial"]))
    if t and t.get("at"):
        return t["at"]
    return clock(r["cs_onset_video_s"] if r["cs_onset_video_s"] is not None
                 else r["us_onset_video_s"])


def is_cr(r):
    return str(r["scored_class"]).startswith("CR")


def is_ur(r):
    return str(r["scored_class"]).startswith("UR")


def is_alpha(r):
    return str(r["scored_class"]).startswith("alpha")


def stim_events(tag):
    """Every pulse read from the LEDs of one recording, accepted or not."""
    f = os.path.join(WORK, tag + "_stim.json")
    if not os.path.exists(f):
        return []
    S = json.load(open(f, encoding="utf-8"))
    out = []
    for key, nm in (("yellow", "CS (yellow LED)"), ("blue", "US (blue LED)")):
        for e in S["events"].get(key, []):
            out.append((e["t"], nm, e))
    out.sort(key=lambda z: z[0])
    return out


# ---------------------------------------------------------------- house style
# The palette and the face come from ebc_config, so a CR is the same blue in a workbook
# chart as in the PNG figures and on the app page.
def F(**kw):
    """A cell font in the house face.  Every Font in this file goes through here."""
    return Font(name=C.FONT, **kw)


INK, MUT, HDR = C.xl("ink"), C.xl("muted"), C.xl("accent")
ACTION, FLAG = C.xl("action"), C.xl("flag_soft")
WHITE = "FFFFFFFF"
thin = Side(style="thin", color=C.xl("rule"))
OUTDIR = OUT

PROTOCOL = [
    ("Design", "%s. A block is %d paired CS-US trials followed by %d CS-only trial(s); "
               "%d blocks of conditioning, then CS-only trials as extinction."
               % (DES["label"], NOM["paired_per_block"], NOM["cs_only_per_block"],
                  NOM["n_blocks"])),
    ("Stimuli", f"CS = yellow LED, {NOM['cs_ms']:.0f} ms. US = blue LED, {NOM['us_dur_ms']:.0f} ms, "
                f"beginning {NOM['us_onset_ms']:.0f} ms after CS onset. {DES['sentence']}"),
    ("CR window", "A blink counted as a conditioned response if it began between "
                   "%.0f and %.0f ms after CS onset. %s"
                   % (CR_LO, CR_HI, WIN["why"])),
    ("Structure recovered", "The trial structure was recovered from the video, not imposed on it: the "
                            "LEDs were read frame by frame and the sequence that came out was then "
                            "compared with the protocol. Recovered %d paired CS-US and %d CS-only trials "
                            "against %d and %d expected; the run of paired trials before each CS-only "
                            "probe was %s. Strict %d+%d x %d structure: %s."
                            % (CHECKS.get("found_paired", 0), CHECKS.get("found_cs_only", 0),
                               CHECKS.get("expected_paired", 0), CHECKS.get("expected_cs_only", 0),
                               CHECKS.get("paired_runs_before_each_probe", []),
                               NOM["paired_per_block"], NOM["cs_only_per_block"], NOM["n_blocks"],
                               "yes" if CHECKS.get("strict_block_structure") else "NO - see the report")),
]
METHOD = [
    ("Stimulus detection", "Both LEDs were recovered from the pixels, never assumed. Each is detected as a "
                           "transient against a running background, so a static coloured object cannot "
                           "trigger it, and the two are detected independently of each other."),
    ("Detection filter", "A genuine CS lasts about %.0f ms. Pulses outside +-%.0f%% of that, or closer "
                         "than %g s to the previous accepted one, are LED flicker rather than stimuli and "
                         "are rejected. Blue pulses that do not match the US duration go the same way. "
                         "Every rejected pulse is listed in stimulus_events.csv, so nothing is dropped "
                         "silently." % (NOM["cs_ms"], NOM["cs_tol"] * 100, NOM["min_iti_s"])),
    ("Trial alignment", "Every trial window is cut with the LED and the face inside the same crop, so the CS "
                        "onset is re-detected inside each window. The alignment error measured for each trial is a column in the trial tables."),
    ("Eyelid measure", "MediaPipe FaceMesh, 478 landmarks with iris refinement, on a 2x-upscaled face crop. "
                       "Eye aspect ratio per eye, averaged. EAR is normalised by eye width, so it is robust "
                       "to head movement and camera distance."),
    ("Closure scale", "0% = a blink-robust open-eye reference (85th percentile of EAR in that window). "
                      "100% = a full-closure reference pooled over every trial of every recording, so "
                      "all blocks sit on one comparable scale. Smoothed with a 5-frame Savitzky-Golay "
                      "filter (42 ms)."),
    ("Blink criterion", "Five robust SDs above the trial's own pre-CS baseline, floor 15% closure, walked "
                        "back along the falling edge to the true onset. A separate blink must re-reach 40% "
                        "closure after first returning below 20%."),
    ("Where the CR window comes from",
     "Neither the eye nor the brainstem responds instantly, so the window is not the bare "
     "interval between the two stimuli. The reflex latency - how long after a puff the "
     "unconditioned blink actually starts - is measured in the US-only baseline as the "
     "mean of those onsets minus %.1f SD, and both edges of the CR window sit that far "
     "after their own stimulus: a blink cannot have been caused by the CS until that long "
     "after CS onset, and cannot have been caused by the US until that long after US "
     "onset. Only trials the scorer stands behind are used, and of those, onsets more "
     "than %.1f robust SDs from the median of that baseline are set aside first - a "
     "mean and an SD have no defence against one spontaneous blink scored long after "
     "the reflex was missed. Every trial set aside is named on the US-only workbook's "
     "cover. %s" % (C.REFLEX_K, C.REFLEX_OUTLIER_SD, WIN["why"])),
    ("Second look after an artefact", "If the first event in the window began before the CR "
                                      "window opens (under %.0f ms, too soon for the CS to have "
                                      "caused it) or the lid was already moving at CS onset, the "
                                      "window is searched for a LATER blink, because a real CR or "
                                      "UR may sit behind the artefact. Where one is found it "
                                      "becomes the scored response; 'First response obscured' marks "
                                      "those trials and the original first blink is kept in its own "
                                      "columns." % CR_LO),
    ("Response classes", f"alpha / startle = onset under {CR_LO:.0f} ms: too soon after the "
                         f"CS for the CS to have caused it. CR = onset {CR_LO:.0f}-{CR_HI:.0f} ms, "
                         f"i.e. the blink was already under way before the puff could have "
                         f"driven it. UR only = onset at or after {CR_HI:.0f} ms. Both boundaries "
                         f"move with the protocol and with the measured reflex, rather than being "
                         f"fixed. Trials where the lid was moving at CS onset and no later blink "
                         f"was found cannot be timed and are excluded from the summary."),
]
CAVEATS = [
    ("Measure", "Landmark EAR is a good proxy for lid aperture but it is not EMG or a magnetic search coil. "
                "Treat the timings as the reliable quantity and absolute closure percentages as relative."),
    ("Which onset to use", "'Scored onset' is the column to analyse - it already applies the second-look "
                           "rule. 'Blink onset' is the raw first event, kept for transparency."),
    ("Quality flags", "Flagged trials are kept in the table so nothing is silently dropped; filter on "
                      "'Quality flag' = clean if you want the strictest subset."),
    ("Trials the scorer will not stand behind",
     "'Score by hand' = yes marks a trial the automatic score cannot be trusted on, with "
     "the reason beside it. They are listed on their own sheet, with the time to open the "
     "recording at, and they are left out of the CR percentages - a trial that cannot be "
     "timed cannot be counted either way. Nothing is dropped: every one is still in the "
     "trial table."),
]

STUDY = CFG["study"]
# The conditioning workbook names the design it was actually run under, from the same
# three numbers everything else reads.
ROLE_TITLE = {"conditioning": DES["label"].lower(),
              "extinction": "extinction / post-conditioning test",
              "baseline_cs": "baseline - CS alone",
              "baseline_us": "baseline - US alone"}
ROLE_FIGS = {"conditioning": [("cond_acquisition.png", "Acquisition by block - conditioned responses "
                                                       "replace reactions to the puff"),
                              ("cond_paired_onset_scatter.png", "Blink onset per paired trial, joined in "
                                                                "order, with the block-mean learning curve"),
                              ("cond_csonly_onset_scatter.png", "The CS-only probe ending each block"),
                              ("cond_paired_overview.png", "Eyelid closure per recording - raster and "
                                                           "overlaid traces")],
             "extinction": [("ext_onset_scatter.png", "Blink onset per CS-only trial, against the "
                                                      "learned US window"),
                            ("ext_overview.png", "Eyelid closure - raster and overlaid traces")],
             "baseline_cs": [("baseline_cs_onset_scatter.png", "Blink onset per CS-only trial"),
                             ("baseline_cs_overview.png", "Eyelid closure - raster and overlaid traces")],
             "baseline_us": [("baseline_us_onset_scatter.png", "Blink latency per US-only trial"),
                             ("baseline_us_overview.png", "Eyelid closure - raster and overlaid traces")]}


def _span(tags):
    labels = [SESSMETA[t]["label"] for t in tags if t in SESSMETA]
    mins = sum(SESSMETA[t]["duration_s"] for t in tags if t in SESSMETA) / 60.0
    return labels, mins


def _book(role):
    tags = [t for t in TAG_ORDER if t in SESSMETA and SESSMETA[t]["role"] == role]
    if not tags or not any(r["role"] == role for r in ALLROWS):
        return None
    labels, mins = _span(tags)
    npair = sum(1 for r in ALLROWS if r["role"] == role and r["trial_type"] == "CS-US")
    nrest = sum(1 for r in ALLROWS if r["role"] == role and r["trial_type"] != "CS-US")
    others = [ROLE_TITLE[o] for o in C.ROLES
              if o != role and any(r["role"] == o for r in ALLROWS)]
    off = M.get("offsets", {})
    timeline = "  ".join("%s = %.0f-%.0f s." % (SESSMETA[t]["label"], off.get(t, 0),
                                                off.get(t, 0) + SESSMETA[t]["duration_s"])
                         for t in tags if t in off)
    what = [("What this covers",
             "%s: %d recording(s) (%s), %.1f min in total. %d paired CS-US trials and %d "
             "CS-only / US-only trials." % (ROLE_TITLE[role], len(labels), ", ".join(labels),
                                            mins, npair, nrest))]
    if role == "conditioning":
        what.append(("CS-only trials are separate",
                     "The CS-only probes are scored but held out of the recording summary, the block "
                     "summary and the main scatter - a trial with no US is a different measurement. "
                     "They have their own sheet and their own figure."))
        what.append(("Protocol check",
                     "%d paired and %d CS-only recovered against %d and %d expected. Run of paired "
                     "trials before each probe: %s. Strict structure: %s."
                     % (CHECKS.get("found_paired", 0), CHECKS.get("found_cs_only", 0),
                        CHECKS.get("expected_paired", 0), CHECKS.get("expected_cs_only", 0),
                        CHECKS.get("paired_runs_before_each_probe", []),
                        "yes" if CHECKS.get("strict_block_structure") else "NO")))
    if role == "extinction":
        what.append(("Why it is separate",
                     "No CS is paired with a US here, so this is extinction, not more conditioning. "
                     "Pooling it with the conditioning trials would mix two measurements."))
        what.append(("Reading the US column",
                     "No US was delivered. 'Closure at US' is sampled at %.0f ms - the interval "
                     "learned during conditioning - and the CR window (%.0f-%.0f ms) is the one "
                     "the conditioning trials were scored against, applied here as a probe: the "
                     "same window, on trials where nothing followed the CS. %s"
                     % (NOM["us_onset_ms"], CR_LO, CR_HI, DES["sentence"])))
    if role == "baseline_us":
        what.append(("What the timings mean",
                     "There is no CS, so every window is anchored on the US and the latencies are "
                     "measured from the puff. These are unconditioned responses by definition and "
                     "give the reflex baseline the CRs are read against."))
        # This workbook is where the CR window is actually measured, so it says which of
        # its own trials went into the number and which did not, and why.
        R = M.get("reflex") or {}
        if R.get("onset_ms") is not None:
            what.append(("The reflex latency measured here",
                         "%d of these trials gave a usable onset: %s ms, mean %.1f, SD %.1f. "
                         "mean - %.1f SD = %.0f ms is the soonest a stimulus can have caused a "
                         "blink, and it is what sets the CR window in the conditioning and "
                         "extinction workbooks: %.0f-%.0f ms after CS onset."
                         % (R["n"], ", ".join("%.0f" % o for o in R.get("onsets", [])),
                            R["mean_ms"], R["sd_ms"], R["k"], R["onset_ms"], CR_LO, CR_HI)))
        else:
            what.append(("The reflex latency could not be measured here",
                         "%s, so the CR window falls back to the protocol's startle cut-off "
                         "and the bare US onset (%.0f-%.0f ms). %s"
                         % (R.get("why", "these trials could not be used"), CR_LO, CR_HI,
                            "Scoring more of these trials by hand would fix it."
                            if R.get("skipped") else "")))
        if R.get("skipped"):
            what.append(("Trials left out of that measurement",
                         "; ".join("%s trial %d (%s)"
                                   % (k["session_name"], k["session_trial"], k["because"])
                                   for k in R["skipped"])
                         + ". A trial whose lid was already moving at the puff has an onset "
                           "that is not the reflex, and one of those would drag the mean out "
                           "and inflate the SD until the window was meaningless."))
    if role == "baseline_cs":
        what.append(("What the timings mean",
                     "The CS alone, before any pairing. Blinks here are orienting or spontaneous, "
                     "not conditioned, and give the false-positive rate for the CR window."))
    nman = sum(1 for r in ALLROWS if r["role"] == role and r["needs_manual_scoring"])
    what.append(("Trials to score by hand",
                 ("%d of the %d trial(s) here could not be scored with confidence. They are "
                  "on the 'Score by hand' sheet with the time to open the recording at and "
                  "the reason for each, and they are marked in the trial table."
                  % (nman, npair + nrest)) if nman else
                 "None - every trial in this workbook was scored cleanly."))
    what.append(("Timeline", (timeline + " 'CS onset on session clock' is the position on that "
                              "continuous clock.") if timeline else "Single recording."))
    if others:
        what.append(("Companion workbooks", "This run also produced: " + ", ".join(others) + "."))
    return dict(file="EBC_%s_%s.xlsx" % (STUDY, role),
                title="%s - %s - %s" % (STUDY, ROLE_TITLE[role], ", ".join(labels)),
                sel=(lambda rl: (lambda r: r["role"] == rl))(role),
                figs=ROLE_FIGS.get(role, []),
                what=what)


BOOKS = dict((role, b) for role, b in
             ((r, _book(r)) for r in C.ROLES) if b)


COLS = [("#", "gidx", "0", 6), ("Block", "block", "0", 7), ("Trial in block", "trial_in_block", "0", 9),
        ("Session", "session_name", None, 10), ("Trial in session", "session_trial", "0", 9),
        ("Type", "trial_type", None, 10),
        ("CS onset in video (s)", "cs_onset_video_s", "0.000", 14),
        ("CS onset on session clock (s)", "cs_onset_block_s", "0.000", 14),
        ("Alignment error (ms)", "alignment_error_ms", "0.00", 11),
        ("Face tracked (%)", "face_tracked_pct", "0.0", 11),
        ("CS duration measured (ms)", "cs_duration_measured_ms", "0.0", 13),
        ("CS timing source", "cs_timing", None, 20),
        ("Block boundary from", "block_closed_by", None, 16),
        ("Quality flag", "quality", None, 24),
        # The scorer's own verdict on whether this trial can be left to it.  Next to the
        # quality flag because that is what a reader is already looking at when deciding
        # whether to trust a row.
        ("Score by hand", "needs_manual_scoring", None, 11),
        ("Score by hand because", "manual_scoring_because", None, 34),
        ("Blinks in 1 s window", "n_full_blinks", "0", 10),
        ("SCORED onset (ms)", "scored_onset_ms", "0.0", 13),
        ("SCORED class", "scored_class", None, 19),
        ("First response obscured", "first_response_obscured", None, 12),
        ("Raw first blink onset (ms)", "blink_onset_ms", "0.0", 13),
        ("Raw first blink class", "response_class", None, 19),
        ("Later blink onset (ms)", "secondary_onset_ms", "0.0", 13),
        ("Later blink peak (%)", "secondary_peak_pct", "0.0", 12),
        ("Later blink class", "secondary_class", None, 19),
        ("Peak closure time (ms)", "peak_closure_ms", "0.0", 13),
        ("Peak closure (%)", "peak_closure_pct", "0.0", 11),
        ("Closing speed (%/ms)", "closing_speed_pct_per_ms", "0.00", 12),
        ("Closure duration (ms)", "closure_duration_ms", "0.0", 13),
        ("Closure at US %.0f ms (%%)" % NOM["us_onset_ms"], "closure_at_US_pct", "0.0", 13),
        ("Closure at CS offset %.0f ms (%%)" % NOM["cs_ms"], "closure_at_CSoff_pct", "0.0", 14),
        ("Half-reopened (ms)", "reopen_half_ms", "0.0", 12),
        ("Fully reopened (ms)", "reopen_full_ms", "0.0", 12),
        ("Closure at 1000 ms (%)", "closure_at_1000ms_pct", "0.0", 13),
        ("Closed at US", "closed_at_US", None, 10),
        ("Reopened before US", "reopened_before_US", None, 11),
        ("All blink onsets (ms)", "all_blink_onsets_ms", None, 16),
        ("All blink amplitudes (%)", "all_blink_amps_pct", None, 16),
        ("Inter-blink interval (ms)", "inter_blink_ms", None, 14),
        ("Partial lid movements", "partial_movement_ms", None, 20)]

# A trace protocol has an interval with neither stimulus on, and how closed the lid was
# across it is a measurement in its own right.  A delay protocol has no such interval, so
# the column is not offered there rather than being filled with blanks.
if DES["kind"] == "trace":
    _mid = (NOM["cs_ms"] + NOM["us_onset_ms"]) / 2
    COLS.insert([c[1] for c in COLS].index("closure_at_CSoff_pct") + 1,
                ("Closure mid-gap %.0f ms (%%)" % _mid, "closure_at_midgap_pct", "0.0", 14))


def col_letter(field):
    """Spreadsheet column letter of a trial-table field, by name.

    The conditional formats used to name their columns as literal letters, which drifts
    the moment a column is added or moved - and the trace column above does exactly that.
    """
    return get_column_letter([c[1] for c in COLS].index(field) + 1)


def style_header(ws, row=1):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = F(bold=True, size=10, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=HDR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 36


def widths(ws, w):
    for i, x in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(i)].width = x


def scoreable(rs):
    return [r for r in rs if r["scored_class"] not in (None, MOVING_LBL)]


def write_table(ws, rows):
    for j, c in enumerate(COLS, 1):
        ws.cell(row=1, column=j, value=c[0])
    for i, src in enumerate(rows, 2):
        flagged = src["quality"] != "clean"
        for j, (hh, key, fmt, w) in enumerate(COLS, 1):
            v = src.get(key)
            if isinstance(v, bool):
                v = "yes" if v else "no"
            c = ws.cell(row=i, column=j, value=v)
            c.font = F(size=10, color=MUT, italic=True) if flagged else F(size=10, color=INK)
            c.alignment = Alignment(horizontal="right" if fmt else "left")
            if fmt:
                c.number_format = fmt
            c.border = Border(bottom=thin)
            if key in ("scored_onset_ms", "scored_class"):
                c.font = F(size=10, bold=True, color=INK, italic=flagged)
            # A trial nobody should read off this row without opening the video is said
            # so on the row itself, in the one warm colour this workbook uses, rather
            # than only on its own sheet.
            if src.get("needs_manual_scoring") and key in ("needs_manual_scoring",
                                                           "manual_scoring_because"):
                c.fill = PatternFill("solid", fgColor=FLAG)
                c.font = F(size=10, color=ACTION,
                           bold=(key == "needs_manual_scoring"))
    style_header(ws)
    widths(ws, [c[3] for c in COLS])
    ws.freeze_panes = "G2"
    n = len(rows) + 1
    if not rows:
        ws.cell(row=2, column=1, value="No trials of this kind in this recording group.")
        ws.cell(row=2, column=1).font = F(size=10, color=MUT, italic=True)
        return
    ws.auto_filter.ref = "A1:" + get_column_letter(len(COLS)) + str(n)
    # latency, coloured across the window this protocol actually uses: nothing before the
    # startle cut-off, the US onset in the middle, the end of the search window at the top
    lat = col_letter("scored_onset_ms")
    ws.conditional_formatting.add("%s2:%s%d" % (lat, lat, n), ColorScaleRule(
        start_type="num", start_value=0, start_color=C.xl("cs_soft"),
        mid_type="num", mid_value=round(CR_HI / 2), mid_color=C.xl("us_soft"),
        end_type="num", end_value=round(max(DES["span_ms"], CR_HI)),
        end_color=C.xl("ur_soft")))
    pct = ["peak_closure_pct", "closure_at_US_pct", "closure_at_CSoff_pct",
           "closure_at_1000ms_pct", "closure_at_midgap_pct"]
    for field in pct:
        if field not in [c[1] for c in COLS]:
            continue
        L = col_letter(field)
        ws.conditional_formatting.add("%s2:%s%d" % (L, L, n), ColorScaleRule(
            start_type="num", start_value=0, start_color=WHITE,
            end_type="num", end_value=100, end_color=C.xl("us_mid")))
    ws.sheet_view.showGridLines = False


def scatter_sheet(ws, rows, has_us, xlabel, title, us_anchored=False):
    # A US-only baseline is anchored on the puff, so its trials carry the US-anchored
    # response labels and time zero is the US.  The CS reference lines mean nothing there
    # and the CS-anchored labels never appear, so both are swapped out rather than left
    # to mislead - and, before this, to raise a KeyError on the first US-only trial.
    if us_anchored:
        hd = ["#", "Block", "Session", UR_PUFF_LBL, ALPHA_US_LBL, "",
              "lid moving at onset", "US onset = 0 ms", "", "", "Block mean latency"]
        colmap = {UR_PUFF_LBL: 4, ALPHA_US_LBL: 5, MOVING_LBL: 7}
        yaxis = "Blink latency (ms from blue LED / US onset)"
        ymin, ymax = -120, 400
    else:
        lbl_us = ("US onset = %.0f ms" % NOM["us_onset_ms"]) if has_us else (
            "learned US onset = %.0f ms (none delivered)" % NOM["us_onset_ms"])
        # The reference line at the CS offset is only "the CS and US offset" when the two
        # really do end together; in a trace protocol it is the start of the gap.
        lbl_cs_off = ("CS / US offset = %.0f ms" % NOM["cs_ms"] if DES["coterminate"]
                      else "CS offset = %.0f ms" % NOM["cs_ms"])
        hd = ["#", "Block", "Session", CR_LBL, ALPHA_LBL, UR_LBL,
              "lid moving at onset", "CS onset = 0 ms", lbl_us,
              lbl_cs_off, "Block mean onset",
              "CR window opens = %.0f ms" % CR_LO, "CR window closes = %.0f ms" % CR_HI]
        colmap = {CR_LBL: 4, ALPHA_LBL: 5, UR_LBL: 6, MOVING_LBL: 7}
        yaxis = "Blink onset (ms from yellow LED / CS onset)"
        # room for the whole trial, so a trace protocol's UR is on the chart rather than
        # clipped off the top of it
        ymin, ymax = -120, max(520, int(DES["span_ms"]) + 160)
    for j, x in enumerate(hd, 1):
        if x:
            ws.cell(row=1, column=j, value=x)
    bmean = {}
    for b in sorted({r["block"] for r in rows if r["block"]}):
        g = [r["scored_onset_ms"] for r in rows if r["block"] == b and r["scored_onset_ms"] is not None
             and r["scored_class"] != MOVING_LBL]
        if g:
            bmean[b] = float(np.mean(g))
    for i, r_ in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=r_["gidx"])
        ws.cell(row=i, column=2, value=r_["block"])
        ws.cell(row=i, column=3, value=r_["session_name"])
        col = colmap.get(r_["scored_class"])
        if r_["scored_onset_ms"] is not None and col:
            ws.cell(row=i, column=col, value=r_["scored_onset_ms"]).number_format = "0.0"
        ws.cell(row=i, column=8, value=0)
        if not us_anchored:
            ws.cell(row=i, column=9, value=NOM["us_onset_ms"])
            ws.cell(row=i, column=10, value=NOM["cs_ms"])
            ws.cell(row=i, column=12, value=CR_LO)
            ws.cell(row=i, column=13, value=CR_HI)
        if r_["block"] in bmean:
            ws.cell(row=i, column=11, value=round(bmean[r_["block"]], 1)).number_format = "0.0"
    style_header(ws)
    widths(ws, [6, 7, 10, 19, 21, 19, 17, 15, 20, 19, 15] + ([] if us_anchored else [19, 19]))
    ws.sheet_view.showGridLines = False
    ch = ScatterChart()
    ch.style = 2
    ch.title = title
    ch.x_axis.title = xlabel
    ch.y_axis.title = yaxis
    ch.height, ch.width = 13, 30 if len(rows) > 14 else 20
    ch.x_axis.scaling.min, ch.x_axis.scaling.max = 0, len(rows) + 1
    ch.y_axis.scaling.min, ch.y_axis.scaling.max = ymin, ymax
    xref = Reference(ws, min_col=1, min_row=2, max_row=len(rows) + 1)
    marks = ((4, C.xl("us"), 7), (5, C.xl("cs"), 7), (7, C.xl("faint"), 6)) if us_anchored else (
             (4, C.xl("cr"), 7), (5, C.xl("cs"), 7), (6, C.xl("ur"), 7), (7, C.xl("faint"), 6))
    for k, colr, size in marks:
        s = Series(Reference(ws, min_col=k, min_row=1, max_row=len(rows) + 1), xref, title_from_data=True)
        s.marker = Marker(symbol="circle", size=size)
        s.marker.graphicalProperties = GraphicalProperties(solidFill=colr)
        s.marker.graphicalProperties.line = LineProperties(solidFill=colr)
        s.graphicalProperties.line.noFill = True
        ch.series.append(s)
    lines = ((8, C.xl("us") if us_anchored else C.xl("cs"), "solid", 20000),
             (11, C.xl("ink"), "solid", 28000)) if us_anchored else (
            [(8, C.xl("cs"), "solid", 20000), (9, C.xl("us"), "dash", 20000),
             (10, C.xl("us"), "dash", 20000), (11, C.xl("ink"), "solid", 28000)]
            # Both edges of the CR window, on the chart rather than in a caption - but
            # only when they were measured, or they would simply redraw the US line.
            + ([(12, C.xl("cr"), "sysDash", 16000), (13, C.xl("cr"), "sysDash", 16000)]
               if WIN["measured"] else []))
    for k, colr, dash, wdt in lines:
        s = Series(Reference(ws, min_col=k, min_row=1, max_row=len(rows) + 1), xref, title_from_data=True)
        s.marker = Marker(symbol="none")
        s.graphicalProperties.line = LineProperties(solidFill=colr, w=wdt, prstDash=dash)
        ch.series.append(s)
    ws.add_chart(ch, "M2")


LOGO = os.path.join(BASE, "assets", "logo_full.png")
LOGO_W = 300                                    # px; the tagline is still legible here


def masthead(ws):
    """Put the lab's mark at the top of a cover sheet.  Returns the first free row.

    A workbook is the thing that actually leaves this folder, so it carries the mark and
    the build that produced it.  If the logo file is not there the sheet is simply built
    without it - a missing image is never a reason to lose a run's numbers.
    """
    r = 1
    try:
        im = XLImage(LOGO)
        im.width, im.height = LOGO_W, round(im.height * LOGO_W / im.width)
        ws.row_dimensions[1].height = im.height * 0.75 + 6      # px -> points, plus air
        ws.add_image(im, "A1")
        r = 2
    except (OSError, ValueError):
        pass
    c = ws.cell(row=r, column=1, value="%s  -  EBC Analyzer %s" % (C.LAB, C.VERSION))
    c.font = F(size=9, color=MUT)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    return r + 2


def build(block):
    cfg = BOOKS[block]
    rows = sorted([r for r in ALLROWS if cfg["sel"](r)], key=lambda r: (r["trial_type"] != "CS-US", r["gidx"]))
    paired = [r for r in rows if r["trial_type"] == "CS-US"]
    csonly = [r for r in rows if r["trial_type"] == "CS-only"]
    # A US-only baseline holds neither paired nor CS-only trials: every row is "US-only",
    # so it used to match neither list and the workbook came out with an empty trial table
    # and an empty scatter, while the CSV and the figures had the trials all along.
    # Whenever there are no paired trials, the single table is everything that is not one.
    main = paired if paired else [r for r in rows if r["trial_type"] != "CS-US"]
    main_type = "Paired" if paired else (main[0]["trial_type"] if main else "CS-only")
    main_sheet = "Paired trials" if paired else "%s trials" % main_type
    sess = sorted({r["session"] for r in rows}, key=lambda t: TAG_ORDER.index(t))
    wb = Workbook()

    def blockhdr(ws, r, title):
        c = ws.cell(row=r, column=1, value=title)
        c.font = F(bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor=HDR)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

    def kv(ws, r, k, v, h=46):
        ws.cell(row=r, column=1, value=k).font = F(size=10, color=MUT)
        c = ws.cell(row=r, column=2, value=v)
        c.font = F(size=10, bold=True, color=INK)
        c.alignment = Alignment(horizontal="left", wrap_text=True)
        ws.row_dimensions[r].height = h

    # ---------------- Read me ----------------
    # These sheets get mailed around and pasted into papers on their own, so the cover
    # says whose tool made them and which build did it, next to what they contain.
    ws = wb.active
    ws.title = "Read me"
    r = masthead(ws)
    ws.cell(row=r, column=1, value=cfg["title"]).font = F(bold=True, size=14, color=INK)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 2
    for hdr_, items in (("PROTOCOL", PROTOCOL), ("WHAT THIS WORKBOOK COVERS", cfg["what"]),
                        ("HOW THE NUMBERS WERE PRODUCED", METHOD), ("CAVEATS", CAVEATS)):
        blockhdr(ws, r, hdr_); r += 1
        for k, v in items:
            kv(ws, r, k, v); r += 1
        r += 1
    blockhdr(ws, r, "SHEETS"); r += 1
    sheets = [("Score by hand", "The trials the scorer will not stand behind, and where "
                                "to find each one in its recording."),
              ("Session summary", "One row per recording, paired CS-US trials only."),
              ("Block summary", "One row per block of 9 paired trials - the learning curve in numbers.")] if paired else [
              ("Score by hand", "The trials the scorer will not stand behind, and where "
                                "to find each one in its recording."),
              ("Trial summary", "Counts, and mean / SD / median of the scored onsets.")]
    sheets += ([("Paired trials", "One row per paired CS-US trial."),
                ("CS-only trials", "The CS-only probes, scored and charted separately.")] if paired
               else [(main_sheet, "One row per %s trial." % main_type.lower())])
    sheets += [
               ("Stimulus events", "Every accepted CS and US event with frame and time."),
               ("Onset scatter", "Data behind the scatter, with a live Excel chart."),
               ("Closure traces", "Full eyelid traces, time x trial, with a live chart."),
               ("Figures", "Rendered PNG figures.")]
    for k, v in sheets:
        kv(ws, r, k, v, 20); r += 1
    widths(ws, [30, 112])
    ws.sheet_view.showGridLines = False

    # ---------------- Score by hand ----------------
    # The trials the scorer will not stand behind, on a sheet of their own and first,
    # because they are the only thing in the workbook that asks somebody to do something.
    # Every one of them is still in the trial table with the same reason beside it; this
    # sheet exists so that "which ones do I have to look at?" is answered by opening the
    # workbook rather than by filtering it.
    ws = wb.create_sheet("Score by hand")
    hd = ["Recording", "Trial in recording", "#", "Block", "At in the recording",
          "On session clock (s)", "Type", "What the scorer put down",
          "Scored onset (ms)", "Why it cannot be trusted"]
    for j, x in enumerate(hd, 1):
        ws.cell(row=1, column=j, value=x)
    hand = [r for r in rows if r["needs_manual_scoring"]]
    # paired conditioning trials first: those are the measurement itself
    hand.sort(key=lambda r: (r["trial_type"] != "CS-US", TAG_ORDER.index(r["session"]),
                             r["session_trial"]))
    ri = 2
    for r in hand:
        vals = [r["session_name"], r["session_trial"], r["gidx"], r["block"],
                at_in_video(r), r.get("session_clock_s"), r["trial_type"],
                r["scored_class"], r["scored_onset_ms"], r["manual_scoring_because"]]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=j, value=v)
            c.font = F(size=10, color=INK if j != 10 else ACTION)
            c.border = Border(bottom=thin)
            c.alignment = Alignment(horizontal="left" if j in (1, 5, 7, 8, 10) else "right",
                                    wrap_text=(j == 10), vertical="top")
            if j == 6:
                c.number_format = "0.000"
            if j == 9:
                c.number_format = "0.0"
        ws.row_dimensions[ri].height = 30
        ri += 1
    if not hand:
        c = ws.cell(row=2, column=1,
                    value="Every trial in this workbook was scored cleanly - there is "
                          "nothing here to read off the video by hand.")
        c.font = F(size=10, color=MUT, italic=True)
        ri = 3
    note = ("The times are from the start of that recording, in the form a video player "
            "shows. Open the recording there and read the blink onset off it. "
            "%d of the %d trial(s) in this workbook are listed. These trials are left out "
            "of the CR percentages - a trial that cannot be timed cannot be counted "
            "either way - but every one of them is still in the trial table, with "
            "'Score by hand' set and the same reason beside it. Nothing has been dropped."
            % (len(hand), len(rows)))
    c = ws.cell(row=ri + 1, column=1, value=note)
    c.font = F(size=9, color=MUT, italic=True)
    ws.merge_cells(start_row=ri + 1, start_column=1, end_row=ri + 1, end_column=10)
    ws.cell(row=ri + 1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[ri + 1].height = 46
    style_header(ws)
    widths(ws, [16, 13, 6, 7, 15, 15, 10, 21, 12, 62])
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    # ---------------- Trial summary (books with no paired trials) ----------------
    # Extinction and the two baselines have no blocks and no learning curve, so they got
    # no summary at all and the mean and SD had to be worked out by hand from the trial
    # table.  One row per recording, plus a pooled row, on the same measures.
    if not paired and main:
        ws = wb.create_sheet("Trial summary")
        hd = ["Session", "Video file", "Duration (s)", "%s trials" % main_type,
              "Clean (unflagged)", "To score by hand", "Scoreable", "With a blink",
              "CR n", "CR % of scoreable",
              # a US-anchored book has no CS to be startled by, so its own cut-off is the
              # one the scorer used there, not the CR window's lower edge
              ("alpha <20 ms" if main_type == "US-only" else "alpha <%.0f ms" % ALPHA),
              "UR / late", "Mean scored onset (ms)", "SD (ms)", "SEM (ms)",
              "Median scored onset (ms)", "Min (ms)", "Max (ms)",
              "Mean peak closure (%)", "SD peak closure (%)",
              "Mean closure at US %d ms (%%)" % NOM["us_onset_ms"]]
        for j, x in enumerate(hd, 1):
            ws.cell(row=1, column=j, value=x)
        ri = 2
        items = [(t, REC_FILE[t], [r for r in main if r["session"] == t]) for t in sess]
        if len(sess) > 1:
            items.append(("ALL", " + ".join(REC_FILE[t] for t in sess), main))
        for tag, vid, rs in items:
            if not rs:
                continue
            sc = scoreable(rs)
            cr = [r for r in sc if is_cr(r)]
            o = [r["scored_onset_ms"] for r in sc if r["scored_onset_ms"] is not None]
            pk = [r["peak_closure_pct"] for r in sc if r["peak_closure_pct"]]
            n = len(o)
            sd = float(np.std(o, ddof=1)) if n > 1 else None
            vals = [("All pooled" if tag == "ALL" else rs[0]["session_name"]), vid,
                    round(sum(SESSMETA[t]["duration_s"] for t in sess), 1) if tag == "ALL"
                    else round(SESSMETA[tag]["duration_s"], 1),
                    len(rs),
                    sum(1 for r in rs if r["quality"] == "clean"),
                    sum(1 for r in rs if r["needs_manual_scoring"]),
                    len(sc),
                    n,
                    len(cr), round(len(cr) / len(sc) * 100, 1) if sc else None,
                    sum(1 for r in sc if is_alpha(r)),
                    sum(1 for r in sc if is_ur(r)),
                    round(float(np.mean(o)), 1) if o else None,
                    round(sd, 1) if sd is not None else None,
                    round(sd / np.sqrt(n), 1) if sd is not None else None,
                    round(float(np.median(o)), 1) if o else None,
                    round(float(np.min(o)), 1) if o else None,
                    round(float(np.max(o)), 1) if o else None,
                    round(float(np.mean(pk)), 1) if pk else None,
                    round(float(np.std(pk, ddof=1)), 1) if len(pk) > 1 else None,
                    round(float(np.mean([r["closure_at_US_pct"] for r in sc])), 1) if sc else None]
            for j, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=j, value=v)
                c.font = F(size=10, bold=(tag == "ALL"), color=INK)
                c.border = Border(bottom=thin,
                                  top=Side(style="medium", color=INK) if tag == "ALL" else None)
                c.alignment = Alignment(horizontal="left" if j <= 2 else "right")
            ri += 1
        # what the columns actually mean, next to the numbers rather than in the Read me
        ws.cell(row=ri + 1, column=1, value=(
            "Scoreable = a trial whose response could be classified; a lid already closing at "
            "stimulus onset cannot be, and is excluded here but kept in the trial table. "
            "Mean / SD / SEM are over the scored onsets of the scoreable trials. "
            "n is small in a baseline recording - read the SD alongside it, not instead of it."
        )).font = F(size=9, color=MUT, italic=True)
        style_header(ws)
        widths(ws, [13, 26, 11] + [12] * (len(hd) - 3))
        ws.freeze_panes = "C2"
        ws.sheet_view.showGridLines = False

    # ---------------- Session summary (paired only) ----------------
    if paired:
        ws = wb.create_sheet("Session summary")
        hd = ["Session", "Video file", "Duration (s)", "Paired CS-US trials",
              "To score by hand", "Scoreable", "CR n",
              "CR % of scoreable", "alpha <%.0f ms" % ALPHA, "UR only", "no later blink",
              "Mean CR onset (ms)",
              "SD (ms)", "Median CR onset (ms)", "Mean peak closure (%)", "Mean closure at US (%)",
              "Recovered behind artefact"]
        for j, x in enumerate(hd, 1):
            ws.cell(row=1, column=j, value=x)
        ri = 2
        items = [(t, REC_FILE[t], [r for r in paired if r["session"] == t]) for t in sess]
        items.append(("ALL", " + ".join(REC_FILE[t] for t in sess), paired))
        for tag, vid, rs in items:
            if not rs:
                continue
            sc = scoreable(rs)
            cr = [r for r in sc if is_cr(r)]
            o = [r["scored_onset_ms"] for r in cr]
            name = "All pooled" if tag == "ALL" else rs[0]["session_name"]
            vals = [name, vid, round(SESSMETA[sess[0]]["duration_s"], 1) if tag != "ALL"
                    else round(sum(SESSMETA[t]["duration_s"] for t in sess), 1),
                    len(rs), sum(1 for r in rs if r["needs_manual_scoring"]),
                    len(sc), len(cr), round(len(cr) / len(sc) * 100, 1) if sc else None,
                    sum(1 for r in sc if is_alpha(r)),
                    sum(1 for r in sc if is_ur(r)),
                    len(rs) - len(sc),
                    round(float(np.mean(o)), 1) if o else None,
                    round(float(np.std(o, ddof=1)), 1) if len(o) > 1 else None,
                    round(float(np.median(o)), 1) if o else None,
                    round(float(np.mean([r["peak_closure_pct"] for r in sc if r["peak_closure_pct"]])), 1) if sc else None,
                    round(float(np.mean([r["closure_at_US_pct"] for r in sc])), 1) if sc else None,
                    sum(1 for r in rs if r["first_response_obscured"] == "yes" and r["secondary_onset_ms"])]
            if tag != "ALL":
                vals[2] = round(SESSMETA[tag]["duration_s"], 1)
            for j, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=j, value=v)
                c.font = F(size=10, bold=(tag == "ALL"), color=INK)
                c.border = Border(bottom=thin,
                                  top=Side(style="medium", color=INK) if tag == "ALL" else None)
                c.alignment = Alignment(horizontal="left" if j <= 2 else "right")
            ri += 1
        style_header(ws)
        widths(ws, [13, 26, 11, 12, 13, 10, 8, 12, 11, 9, 11, 13, 9, 13, 13, 14, 13])
        ws.freeze_panes = "C2"
        ws.sheet_view.showGridLines = False

        # ---------------- Block summary ----------------
        ws = wb.create_sheet("Block summary")
        hd = ["Block", "Paired trials", "To score by hand", "Scoreable", "CR n", "CR %",
              "UR only n", "UR only %", "alpha n", "Mean scored onset (ms)", "SD (ms)",
              "Mean CR onset (ms)", "Mean peak closure (%)", "Mean closure at US (%)"]
        for j, x in enumerate(hd, 1):
            ws.cell(row=1, column=j, value=x)
        for bi, b in enumerate(sorted({r["block"] for r in paired}), 2):
            g = [r for r in paired if r["block"] == b]
            sc = scoreable(g)
            cr = [r for r in sc if is_cr(r)]
            allo = [r["scored_onset_ms"] for r in sc]
            o = [r["scored_onset_ms"] for r in cr]
            vals = [b, len(g), sum(1 for r in g if r["needs_manual_scoring"]),
                    len(sc), len(cr), round(len(cr) / len(sc) * 100, 1) if sc else None,
                    sum(1 for r in sc if is_ur(r)),
                    round(100 * sum(1 for r in sc if is_ur(r)) / len(sc), 1) if sc else None,
                    sum(1 for r in sc if is_alpha(r)),
                    round(float(np.mean(allo)), 1) if allo else None,
                    round(float(np.std(allo, ddof=1)), 1) if len(allo) > 1 else None,
                    round(float(np.mean(o)), 1) if o else None,
                    round(float(np.mean([r["peak_closure_pct"] for r in sc if r["peak_closure_pct"]])), 1) if sc else None,
                    round(float(np.mean([r["closure_at_US_pct"] for r in sc])), 1) if sc else None]
            for j, v in enumerate(vals, 1):
                c = ws.cell(row=bi, column=j, value=v)
                c.font = F(size=10, color=INK)
                c.border = Border(bottom=thin)
                c.alignment = Alignment(horizontal="right")
        style_header(ws)
        widths(ws, [8, 11, 13, 10, 8, 9, 10, 10, 8, 15, 9, 15, 14, 15])
        ws.sheet_view.showGridLines = False
        n = len(set(r["block"] for r in paired)) + 1
        ws.conditional_formatting.add("F2:F" + str(n), ColorScaleRule(
            start_type="num", start_value=0, start_color=WHITE,
            end_type="num", end_value=100, end_color=C.xl("us_mid")))
        lc = LineChart()
        lc.title = "Learning across blocks"
        lc.x_axis.title = "Block"
        lc.y_axis.title = "CR % / mean onset (ms)"
        lc.height, lc.width = 10, 20
        lc.add_data(Reference(ws, min_col=6, min_row=1, max_row=n), titles_from_data=True)
        lc.add_data(Reference(ws, min_col=12, min_row=1, max_row=n), titles_from_data=True)
        lc.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n))
        ws.add_chart(lc, "P2")

    # ---------------- trial tables ----------------
    write_table(wb.create_sheet(main_sheet), main)
    if paired and csonly:
        write_table(wb.create_sheet("CS-only trials"), csonly)

    # ---------------- Stimulus events ----------------
    ws = wb.create_sheet("Stimulus events")
    hd = ["Recording", "Event #", "Stimulus", "Onset frame", "Onset (s)", "Duration (ms)",
          "Accepted", "Rejected because"]
    for j, x in enumerate(hd, 1):
        ws.cell(row=1, column=j, value=x)
    ri = 2
    for tag in sess:
        name = SESSMETA[tag]["label"]
        for i, (t, kind, e) in enumerate(stim_events(tag), 1):
            vals = [name, i, kind, e["frame"], round(e["t"], 4), e["dur_ms"],
                    "yes" if e["ok"] else "no",
                    e.get("reason", "") or ("" if e["ok"] else "duration off-spec")]
            for j, v in enumerate(vals, 1):
                cc = ws.cell(row=ri, column=j, value=v)
                cc.font = F(size=10, color=INK if e["ok"] else MUT,
                               italic=not e["ok"])
                cc.border = Border(bottom=thin)
                cc.alignment = Alignment(horizontal="left" if j in (1, 3, 7, 8) else "right")
                if j == 5:
                    cc.number_format = "0.0000"
                if j == 6:
                    cc.number_format = "0.0"
            ri += 1
    style_header(ws)
    widths(ws, [22, 9, 17, 12, 12, 13, 10, 26])
    ws.freeze_panes = "C2"
    ws.sheet_view.showGridLines = False

    # ---------------- Onset scatter ----------------
    us_anchored = bool(main) and main[0]["trial_type"] == "US-only"
    scatter_sheet(wb.create_sheet("Onset scatter"), main, bool(paired),
                  "Paired CS-US trial, in order" if paired else "%s trial" % main_type,
                  "Blink latency per trial, from the puff" if us_anchored
                  else "Blink onset per trial, relative to CS and US", us_anchored)
    if paired and csonly:
        scatter_sheet(wb.create_sheet("CS-only scatter"), csonly, False,
                      "CS-only probe (one per block)", "CS-only probes - blink onset")

    # ---------------- Closure traces ----------------
    ws = wb.create_sheet("Closure traces")
    t = None
    series = []
    for r in rows:
        TRC = M["traces"][r["session"]][str(r["session_trial"])]
        if t is None:
            t = TRC["t"]
        series.append((f"{r['session_name']} T{r['session_trial']} ({r['trial_type']})", TRC["C"]))
    ws.cell(row=1, column=1, value="Time from CS onset (ms)")
    for j, (nm, _) in enumerate(series, 2):
        ws.cell(row=1, column=j, value=nm)
    for i, tv in enumerate(t, 2):
        ws.cell(row=i, column=1, value=round(tv, 2)).number_format = "0.00"
        for j, (_, trace) in enumerate(series, 2):
            if i - 2 < len(trace):
                ws.cell(row=i, column=j,
                        value=round(trace[i - 2] * 100, 2)).number_format = "0.0"
    style_header(ws)
    widths(ws, [21] + [17] * len(series))
    ws.freeze_panes = "B2"
    ws.sheet_view.showGridLines = False
    lc = LineChart()
    lc.title = "Eyelid closure (%) - every trial, aligned to yellow LED onset"
    lc.x_axis.title = "Time from CS onset (ms)"
    lc.y_axis.title = "% eyelid closure"
    lc.height, lc.width = 12, 28
    lc.add_data(Reference(ws, min_col=2, max_col=len(series) + 1, min_row=1, max_row=len(t) + 1),
                titles_from_data=True)
    lc.set_categories(Reference(ws, min_col=1, min_row=2, max_row=len(t) + 1))
    for s in lc.series:
        s.graphicalProperties.line = LineProperties(solidFill=C.xl("trace"), w=6000)
        s.smooth = False
    lc.x_axis.tickLblSkip = 12
    lc.legend = None
    ws.add_chart(lc, get_column_letter(len(series) + 3) + "2")

    # ---------------- Figures ----------------
    ws = wb.create_sheet("Figures")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Rendered figures"
    ws["A1"].font = F(bold=True, size=13, color=INK)
    r = 3
    for f, cap in cfg["figs"]:
        f = os.path.join(OUT, f)
        if not os.path.exists(f):
            continue
        im = XLImage(f)
        ws.cell(row=r, column=1, value=cap).font = F(bold=True, size=10, color=MUT)
        r += 1
        k = min(1200 / im.width, 1.0)
        im.width, im.height = int(im.width * k), int(im.height * k)
        ws.add_image(im, "A" + str(r))
        r += int(im.height / 19) + 3
    widths(ws, [115])

    out = os.path.join(OUTDIR, cfg["file"])
    try:
        wb.save(out)
    except PermissionError:
        alt = out.replace(".xlsx", "_NEW.xlsx")
        wb.save(alt)
        print(f"!! {os.path.basename(out)} is open in Excel - wrote {os.path.basename(alt)} instead")
        out = alt
    print(f"saved {out}  ({len(paired)} paired + {len(csonly)} CS-only)")


for b in (sys.argv[2:] or list(BOOKS)):
    if b in BOOKS:
        build(b)
